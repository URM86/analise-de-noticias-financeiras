"""
============================================================
src/storage.py
Módulo de persistência de dados.

Responsabilidades:
  1. Salvar artigos analisados em CSV (append mode)
  2. Exportar relatório agregado para Excel (.xlsx)
  3. Carregar histórico de análises anteriores
  4. Deduplicar artigos já processados (evita re-análise)
============================================================
"""

import logging
import os
from datetime import datetime
from pathlib import Path
from typing import Optional

import pandas as pd

from config.settings import ARQUIVO_CSV, ARQUIVO_EXCEL, PASTA_DADOS

logger = logging.getLogger(__name__)

# ── Colunas do CSV de resultados brutos ──────────────────────────────────────
COLUNAS_CSV = [
    "timestamp_coleta",   # quando foi coletado/analisado
    "ticker",
    "id",
    "titulo",
    "descricao",
    "url",
    "fonte",
    "data_pub",
    "termo_busca",
    "sentimento_pt",
    "sentimento_en",
    "confianca",
    "score_positive",
    "score_negative",
    "score_neutral",
    "texto_analisado",
    "erro",
]


def _garantir_pasta(caminho: str) -> None:
    """Cria a pasta se não existir."""
    Path(caminho).mkdir(parents=True, exist_ok=True)


def _extrair_scores(artigo: dict) -> tuple[float, float, float]:
    """Extrai scores individuais do dicionário scores_todos."""
    todos = artigo.get("scores_todos", {})
    return (
        round(todos.get("positive", 0.0), 4),
        round(todos.get("negative", 0.0), 4),
        round(todos.get("neutral",  0.0), 4),
    )


def salvar_artigos_csv(artigos: list[dict]) -> int:
    """
    Salva/adiciona artigos analisados ao CSV principal.

    Usa append mode com header condicional (escreve header
    apenas se o arquivo não existir ainda).

    Retorna o número de artigos salvos.
    """
    if not artigos:
        logger.warning("Nenhum artigo para salvar.")
        return 0

    _garantir_pasta(PASTA_DADOS)

    agora = datetime.now().isoformat(timespec="seconds")
    linhas = []

    for a in artigos:
        pos, neg, neu = _extrair_scores(a)
        linhas.append(
            {
                "timestamp_coleta": agora,
                "ticker":           a.get("ticker", ""),
                "id":               a.get("id", ""),
                "titulo":           a.get("titulo", ""),
                "descricao":        a.get("descricao", "")[:500],  # limita tamanho
                "url":              a.get("url", ""),
                "fonte":            a.get("fonte", ""),
                "data_pub":         a.get("data_pub", ""),
                "termo_busca":      a.get("termo_busca", ""),
                "sentimento_pt":    a.get("sentimento_pt", ""),
                "sentimento_en":    a.get("sentimento_en", ""),
                "confianca":        a.get("confianca", 0.0),
                "score_positive":   pos,
                "score_negative":   neg,
                "score_neutral":    neu,
                "texto_analisado":  a.get("texto_analisado", "")[:300],
                "erro":             a.get("erro", ""),
            }
        )

    df_novo = pd.DataFrame(linhas, columns=COLUNAS_CSV)

    arquivo = Path(ARQUIVO_CSV)
    modo  = "a" if arquivo.exists() else "w"
    header = not arquivo.exists()

    df_novo.to_csv(arquivo, mode=modo, header=header, index=False, encoding="utf-8-sig")
    logger.info(f"✔ {len(linhas)} artigo(s) salvos em '{ARQUIVO_CSV}'.")
    return len(linhas)


def carregar_ids_processados() -> set[str]:
    """
    Lê o CSV e retorna o conjunto de IDs já processados.
    Usado para evitar re-análise de artigos já vistos.
    """
    arquivo = Path(ARQUIVO_CSV)
    if not arquivo.exists():
        return set()

    try:
        df = pd.read_csv(arquivo, usecols=["id"], encoding="utf-8-sig")
        ids = set(df["id"].dropna().astype(str).tolist())
        logger.debug(f"{len(ids)} artigos já processados no histórico.")
        return ids
    except Exception as exc:
        logger.warning(f"Não foi possível carregar IDs processados: {exc}")
        return set()


def exportar_excel(resumos: list[dict]) -> None:
    """
    Gera um relatório Excel com:
    - Sheet 'Resumo': uma linha por ticker com sentimento geral
    - Sheet 'Histórico': todos os artigos analisados até agora
    - Sheet 'Pivot': tabela dinâmica sentimento × ticker

    O Excel usa pandas + openpyxl.
    """
    _garantir_pasta(PASTA_DADOS)

    try:
        with pd.ExcelWriter(ARQUIVO_EXCEL, engine="openpyxl") as writer:
            # ── Sheet 1: Resumo ──────────────────────────────────────────────
            if resumos:
                df_resumo = pd.DataFrame(resumos)
                df_resumo = df_resumo.rename(
                    columns={
                        "ticker":           "Ticker",
                        "sentimento_geral": "Sentimento Geral",
                        "total":            "Total Notícias",
                        "positivos":        "Positivas",
                        "negativos":        "Negativas",
                        "neutros":          "Neutras",
                        "inconclusivos":    "Inconclusivas",
                        "score_medio":      "Confiança Média",
                        "percentual_pos":   "% Positivas",
                        "percentual_neg":   "% Negativas",
                    }
                )
                df_resumo.to_excel(writer, sheet_name="Resumo", index=False)

                # Aplica cores ao sentimento geral
                ws = writer.sheets["Resumo"]
                _colorir_coluna_sentimento(ws, df_resumo, "Sentimento Geral")

            # ── Sheet 2: Histórico completo ──────────────────────────────────
            arquivo_csv = Path(ARQUIVO_CSV)
            if arquivo_csv.exists():
                df_hist = pd.read_csv(arquivo_csv, encoding="utf-8-sig")
                df_hist.to_excel(writer, sheet_name="Histórico", index=False)

                # ── Sheet 3: Pivot ───────────────────────────────────────────
                try:
                    df_pivot = df_hist.pivot_table(
                        index="ticker",
                        columns="sentimento_pt",
                        values="id",
                        aggfunc="count",
                        fill_value=0,
                    )
                    df_pivot.to_excel(writer, sheet_name="Pivot")
                except Exception as e:
                    logger.debug(f"Pivot não gerada: {e}")

        logger.info(f"✔ Relatório Excel salvo em '{ARQUIVO_EXCEL}'.")

    except Exception as exc:
        logger.error(f"Erro ao exportar Excel: {exc}", exc_info=True)


def _colorir_coluna_sentimento(ws, df, coluna: str) -> None:
    """
    Aplica cores de preenchimento às células da coluna de sentimento no Excel.
    Verde → POSITIVO, Vermelho → NEGATIVO, Cinza → NEUTRO/INCONCLUSIVO
    """
    try:
        from openpyxl.styles import PatternFill

        cores = {
            "POSITIVO":    "C6EFCE",  # verde claro
            "NEGATIVO":    "FFC7CE",  # vermelho claro
            "NEUTRO":      "D9D9D9",  # cinza
            "INCONCLUSIVO": "FFEB9C", # amarelo
        }

        col_idx = df.columns.get_loc(coluna) + 1  # openpyxl é 1-based

        for row_idx, valor in enumerate(df[coluna], start=2):  # pula header
            cor = cores.get(str(valor).upper(), "FFFFFF")
            fill = PatternFill(
                start_color=cor, end_color=cor, fill_type="solid"
            )
            ws.cell(row=row_idx, column=col_idx).fill = fill

    except Exception as exc:
        logger.debug(f"Não foi possível colorir células: {exc}")


def carregar_historico() -> pd.DataFrame:
    """
    Carrega todo o histórico de análises do CSV.
    Retorna DataFrame vazio se o arquivo não existir.
    """
    arquivo = Path(ARQUIVO_CSV)
    if not arquivo.exists():
        return pd.DataFrame(columns=COLUNAS_CSV)

    try:
        df = pd.read_csv(arquivo, encoding="utf-8-sig")
        logger.info(f"Histórico carregado: {len(df)} registros.")
        return df
    except Exception as exc:
        logger.error(f"Erro ao carregar histórico: {exc}")
        return pd.DataFrame(columns=COLUNAS_CSV)
